# attendance_system/app.py

from flask import Flask, render_template, request, jsonify, Response, send_from_directory
from flask_cors import CORS
import base64
import binascii
import io
import os
import json
import shutil
import time
from datetime import datetime
import logging
import uuid

import numpy as np
import face_recognition
from deepface import DeepFace

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Import local modules
try:
    from . import database
    from .face_camera import camera, guided_enrollment
except ImportError:
    import database
    from face_camera import camera, guided_enrollment

app = Flask(__name__)
CORS(app)

# Increase max request payload size to 64MB for multi-photo uploads
app.config['MAX_CONTENT_LENGTH'] = 64 * 1024 * 1024

# Configure directories
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads', 'faces')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Initialize database on boot
database.init_db()

ALLOWED_IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.webp'}

# --- Single Page SPA Routes ---
@app.route('/')
@app.route('/dashboard')
@app.route('/students')
@app.route('/attendance')
@app.route('/records')
def index():
    """Serve the single page application"""
    return render_template('index.html')

# --- Serving Uploaded Images ---
@app.route('/uploads/faces/<path:filename>')
def serve_face_image(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# --- Student Management REST APIs ---
@app.route('/api/students', methods=['GET'])
def api_list_students():
    students = database.get_all_students()
    return jsonify(students)

def _student_folder(student_id):
    return os.path.join(app.config['UPLOAD_FOLDER'], student_id)

def _relative_face_path(student_id, filename):
    return os.path.join(student_id, filename).replace('\\', '/')

def _sanitize_pose_label(label):
    if not label:
        return None
    cleaned = ''.join(ch.lower() if ch.isalnum() else '-' for ch in str(label).strip())
    cleaned = '-'.join(part for part in cleaned.split('-') if part)
    return cleaned or None

def _decode_data_url(data_url):
    if not data_url or ',' not in data_url:
        raise ValueError('Invalid camera image payload')

    header, encoded = data_url.split(',', 1)
    try:
        image_bytes = base64.b64decode(encoded)
    except (ValueError, binascii.Error) as exc:
        raise ValueError('Invalid base64 image payload') from exc

    ext = '.png'
    if 'image/jpeg' in header:
        ext = '.jpg'
    elif 'image/webp' in header:
        ext = '.webp'

    return image_bytes, ext

def _load_image_from_data_url(data_url):
    image_bytes, _ = _decode_data_url(data_url)
    return face_recognition.load_image_file(io.BytesIO(image_bytes))

import math

def _estimate_pose_label_from_image(image, target_pose=None):
    face_locations = face_recognition.face_locations(image, model='hog')
    landmarks_list = face_recognition.face_landmarks(image, face_locations=face_locations)

    if not face_locations or not landmarks_list:
        return {
            'face_detected': False,
            'pose_label': None,
            'pose_score': 0.0,
            'matches_target': False,
        }

    landmarks = landmarks_list[0]
    left_eye = np.mean(np.asarray(landmarks.get('left_eye', []), dtype=np.float32), axis=0) if landmarks.get('left_eye') else None
    right_eye = np.mean(np.asarray(landmarks.get('right_eye', []), dtype=np.float32), axis=0) if landmarks.get('right_eye') else None
    nose_bridge = np.mean(np.asarray(landmarks.get('nose_bridge', []), dtype=np.float32), axis=0) if landmarks.get('nose_bridge') else None

    if left_eye is None or right_eye is None or nose_bridge is None:
        return {
            'face_detected': True,
            'pose_label': 'front',
            'pose_score': 0.0,
            'matches_target': target_pose in ('front', 'front-variant') if target_pose else True,
        }

    eye_center = (left_eye + right_eye) / 2.0
    eye_distance = float(np.linalg.norm(right_eye - left_eye)) or 1.0
    yaw = float((nose_bridge[0] - eye_center[0]) / eye_distance)
    pitch = float((nose_bridge[1] - eye_center[1]) / eye_distance)

    # Roll (head tilt): angle between eyes
    dy = float(right_eye[1] - left_eye[1])
    dx = float(right_eye[0] - left_eye[0])
    roll_deg = math.degrees(math.atan2(dy, dx))

    # Smile ratio: mouth width / eye distance
    top_lip = landmarks.get('top_lip', [])
    is_smiling = False
    smile_ratio = 0.0
    if len(top_lip) >= 7:
        mouth_left = np.asarray(top_lip[0], dtype=np.float32)
        mouth_right = np.asarray(top_lip[6], dtype=np.float32)
        mouth_width = float(np.linalg.norm(mouth_right - mouth_left))
        smile_ratio = mouth_width / eye_distance
        if smile_ratio > 0.46:
            is_smiling = True

    score = max(abs(yaw), abs(pitch), abs(roll_deg) / 30.0)

    # Broad pose label determination
    if abs(yaw) >= 0.06:
        pose_label = 'left' if yaw < 0 else 'right'
    elif pitch <= -0.06:
        pose_label = 'up'
    elif pitch >= 0.10:
        pose_label = 'down'
    else:
        pose_label = 'front'

    # Target pose matching evaluation
    matches_target = False
    if target_pose:
        tp = target_pose.lower()
        if tp == 'front':
            matches_target = (abs(yaw) < 0.10 and abs(pitch) < 0.10 and abs(roll_deg) < 12.0)
        elif tp == 'left':
            matches_target = (yaw < -0.05 and abs(pitch) < 0.18)
        elif tp == 'right':
            matches_target = (yaw > 0.05 and abs(pitch) < 0.18)
        elif tp == 'up':
            matches_target = (pitch <= -0.05 and abs(yaw) < 0.16)
        elif tp == 'down':
            matches_target = (pitch >= 0.08 and abs(yaw) < 0.16)
        else:
            matches_target = (pose_label == tp)
    else:
        matches_target = True

    return {
        'face_detected': True,
        'pose_label': pose_label,
        'pose_score': round(score, 3),
        'yaw': round(yaw, 3),
        'pitch': round(pitch, 3),
        'roll': round(roll_deg, 1),
        'smile_ratio': round(smile_ratio, 3),
        'matches_target': matches_target,
    }

def _extract_embedding(image_path):
    """Extract face embedding using DeepFace with fallback to face_recognition"""
    
    # Try DeepFace first (better accuracy)
    try:
        result = DeepFace.represent(
            img_path=image_path,
            model_name='Facenet512',
            enforce_detection=False,
            detector_backend='opencv'
        )
        
        if isinstance(result, list) and result:
            result = result[0]
        if isinstance(result, dict) and 'embedding' in result:
            return np.asarray(result['embedding'], dtype=np.float32)
            
    except Exception as e:
        logger.warning(f"DeepFace embedding failed: {e}")
    
    # Fallback to face_recognition
    try:
        image = face_recognition.load_image_file(image_path)
        encodings = face_recognition.face_encodings(image)
        if encodings:
            return np.asarray(encodings[0], dtype=np.float32)
    except Exception as e:
        logger.warning(f"face_recognition embedding failed: {e}")
    
    return None

def _normalize_embedding(embedding_vectors):
    if not embedding_vectors:
        return None

    stacked = np.vstack(embedding_vectors)
    averaged = stacked.mean(axis=0)
    norm = np.linalg.norm(averaged)
    if norm > 0:
      averaged = averaged / norm
    return averaged.astype(float).tolist()

def _collect_student_media(student_id, uploaded_files, camera_payloads, camera_pose_labels=None):
    student_dir = _student_folder(student_id)
    os.makedirs(student_dir, exist_ok=True)

    stored_relative_paths = []
    embeddings = []
    image_index = 0

    uploads = [item for item in uploaded_files if item and item.filename]
    cameras = [item for item in camera_payloads if item]
    camera_labels = camera_pose_labels or []
    combined_items = [('upload', item) for item in uploads] + [('camera', item) for item in cameras]

    DEFAULT_POSES = [
        'front_neutral', 'front_smiling',
        'slight_left', 'slight_right',
        'medium_left', 'medium_right',
        'tilt_up', 'tilt_down',
        'head_tilt_left', 'head_tilt_right',
    ]

    if len(combined_items) > 20:
        raise ValueError('A maximum of 20 photos is allowed')

    for source, item in combined_items:
        label = DEFAULT_POSES[image_index] if image_index < len(DEFAULT_POSES) else f'photo-{image_index + 1}'
        if source == 'camera' and image_index - len(uploads) < len(camera_labels):
            pose_label = _sanitize_pose_label(camera_labels[image_index - len(uploads)])
            if pose_label:
                label = pose_label

        if source == 'upload':
            original_ext = os.path.splitext(item.filename)[1].lower()
            ext = original_ext if original_ext in ALLOWED_IMAGE_EXTENSIONS else '.jpg'
            filename = f'{label}_{image_index + 1}{ext}'
            absolute_path = os.path.join(student_dir, filename)
            item.save(absolute_path)
        else:
            image_bytes, ext = _decode_data_url(item)
            filename = f'{label}_{image_index + 1}{ext}'
            absolute_path = os.path.join(student_dir, filename)
            with open(absolute_path, 'wb') as image_file:
                image_file.write(image_bytes)

        embedding = _extract_embedding(absolute_path)
        if embedding is not None:
            embeddings.append(embedding)

        stored_relative_paths.append(_relative_face_path(student_id, filename))
        image_index += 1

    if len(stored_relative_paths) < 3:
        raise ValueError('Please provide at least 3 photos')

    averaged_embedding = _normalize_embedding(embeddings)
    if averaged_embedding is None:
        raise ValueError('No detectable face found in the uploaded images')

    return stored_relative_paths, averaged_embedding

def _delete_student_folder(student_id):
    student_dir = _student_folder(student_id)
    if os.path.exists(student_dir):
        shutil.rmtree(student_dir, ignore_errors=True)

def _camera_images_from_request():
    camera_images = request.form.getlist('camera_images[]')
    if not camera_images and request.form.get('camera_image'):
        camera_images = [request.form.get('camera_image')]
    return camera_images

def _camera_pose_labels_from_request():
    pose_labels = request.form.getlist('camera_pose_labels[]')
    if not pose_labels and request.form.get('camera_pose_label'):
        pose_labels = [request.form.get('camera_pose_label')]
    return pose_labels

def _save_student_submission(student_id, name, course, year, semester=None, is_update=False):
    try:
        if not student_id or not name or not course or not year:
            return jsonify({'status': 'error', 'error_message': 'Missing required fields'}), 400

        uploaded_files = request.files.getlist('photo_files[]')
        if not uploaded_files and 'photo' in request.files:
            uploaded_files = [request.files['photo']]

        camera_images = _camera_images_from_request()
        camera_pose_labels = _camera_pose_labels_from_request()

        if not is_update and database.get_student(student_id):
            return jsonify({'status': 'error', 'error_message': f'Student ID "{student_id}" already exists'}), 400

        existing_student = database.get_student(student_id) if is_update else None
        has_new_media = any(file_item and file_item.filename for file_item in uploaded_files) or len(camera_images) > 0

        if not is_update and not has_new_media:
            return jsonify({'status': 'error', 'error_message': 'At least 3 photos are required'}), 400

        if is_update and not has_new_media:
            success, msg = database.update_student(student_id, name, course, year, semester=semester)
            if success:
                if camera.is_running:
                    camera.load_known_faces()
                return jsonify({'status': 'success', 'message': msg})
            return jsonify({'status': 'error', 'error_message': msg}), 400

        if len([file_item for file_item in uploaded_files if file_item and file_item.filename]) + len(camera_images) < 3:
            return jsonify({'status': 'error', 'error_message': 'Please provide at least 3 photos'}), 400

        if is_update:
            _delete_student_folder(student_id)

        stored_paths, averaged_embedding = _collect_student_media(student_id, uploaded_files, camera_images, camera_pose_labels)
        primary_photo = stored_paths[0] if stored_paths else None
        face_images = ','.join(stored_paths)
        face_embedding = json.dumps(averaged_embedding)

        if is_update:
            success, msg = database.update_student(
                student_id,
                name,
                course,
                year,
                semester=semester,
                photo_path=primary_photo,
                face_images=face_images,
                face_embedding=face_embedding,
            )
        else:
            success, msg = database.add_student(
                student_id,
                name,
                course,
                year,
                semester=semester,
                photo_path=primary_photo,
                face_images=face_images,
                face_embedding=face_embedding,
            )

        if success:
            if camera.is_running:
                camera.load_known_faces()
            return jsonify({'status': 'success', 'message': msg})

        _delete_student_folder(student_id)
        return jsonify({'status': 'error', 'error_message': msg}), 400
    except Exception as e:
        _delete_student_folder(student_id)
        return jsonify({'status': 'error', 'error_message': str(e)}), 500

@app.route('/api/add_student', methods=['POST'])
def api_add_student():
    student_id = request.form.get('id')
    name = request.form.get('name')
    course = request.form.get('course')
    year = request.form.get('year')
    semester = request.form.get('semester') or None
    return _save_student_submission(student_id, name, course, year, semester=semester, is_update=False)

@app.route('/api/students/<student_id>', methods=['PUT'])
def api_edit_student(student_id):
    try:
        student = database.get_student(student_id)
        if not student:
            return jsonify({'status': 'error', 'error_message': 'Student not found'}), 404

        name = request.form.get('name')
        course = request.form.get('course')
        year = request.form.get('year')
        semester = request.form.get('semester') or None

        return _save_student_submission(student_id, name, course, year, semester=semester, is_update=True)
    except Exception as e:
        return jsonify({'status': 'error', 'error_message': str(e)}), 500

@app.route('/api/students/<student_id>', methods=['DELETE'])
def api_delete_student(student_id):
    try:
        success, result = database.delete_student(student_id)
        if success:
            _delete_student_folder(student_id)
            if camera.is_running:
                camera.load_known_faces()
            return jsonify({'status': 'success', 'message': 'Student deleted successfully'})
        else:
            return jsonify({'status': 'error', 'error_message': result}), 400
    except Exception as e:
        return jsonify({'status': 'error', 'error_message': str(e)}), 500

@app.route('/api/student/pose-check', methods=['POST'])
def api_student_pose_check():
    try:
        payload = request.get_json(silent=True) or {}
        frame_data = payload.get('frame') or request.form.get('frame')
        target_pose = _sanitize_pose_label(payload.get('target_pose') or request.form.get('target_pose'))

        if not frame_data:
            return jsonify({'status': 'error', 'error_message': 'Missing frame data'}), 400

        image = _load_image_from_data_url(frame_data)
        pose_result = _estimate_pose_label_from_image(image, target_pose=target_pose)

        return jsonify({
            'status': 'success',
            'face_detected': pose_result['face_detected'],
            'pose_label': pose_result['pose_label'],
            'pose_score': pose_result['pose_score'],
            'target_pose': target_pose,
            'matches_target': pose_result['matches_target'],
            'yaw': pose_result.get('yaw', 0),
            'pitch': pose_result.get('pitch', 0),
            'roll': pose_result.get('roll', 0),
            'smile_ratio': pose_result.get('smile_ratio', 0),
        })
    except Exception as e:
        logger.warning(f'Pose check failed: {e}')
        return jsonify({'status': 'error', 'error_message': str(e)}), 500

# --- OpenCV Guided Video Enrollment Routes ---
@app.route('/api/enrollment/start', methods=['POST'])
def api_enrollment_start():
    try:
        data = request.get_json(silent=True) or request.form
        student_id = data.get('id')
        name = data.get('name')
        course = data.get('course')
        year = data.get('year')
        semester = data.get('semester')
        is_update = bool(data.get('is_update'))

        if not student_id or not name or not course or not year:
            return jsonify({'status': 'error', 'error_message': 'Please fill in Student ID, Name, Subject, and Year first.'}), 400

        success = guided_enrollment.start(student_id, name, course, year, semester=semester, is_update=is_update)
        if success:
            return jsonify({'status': 'success', 'message': f'OpenCV live enrollment started for {name}'})
        return jsonify({'status': 'error', 'error_message': guided_enrollment.status_message}), 500
    except Exception as e:
        return jsonify({'status': 'error', 'error_message': str(e)}), 500

def gen_enrollment_frames():
    while True:
        frame_bytes = guided_enrollment.get_frame()
        if frame_bytes is None:
            time.sleep(0.03)
            continue
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')

@app.route('/api/enrollment/video_feed')
def api_enrollment_video_feed():
    return Response(gen_enrollment_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/api/enrollment/status', methods=['GET'])
def api_enrollment_status():
    return jsonify({'status': 'success', 'enrollment': guided_enrollment.get_status()})

@app.route('/api/enrollment/stop', methods=['POST'])
def api_enrollment_stop():
    guided_enrollment.stop()
    return jsonify({'status': 'success', 'message': 'OpenCV enrollment stream stopped'})

@app.route('/api/enrollment/capture', methods=['POST'])
def api_enrollment_capture():
    """Trigger an immediate manual capture for the current pose step."""
    ok = guided_enrollment.trigger_manual_capture()
    if ok:
        return jsonify({'status': 'success', 'message': 'Manual capture triggered'})
    return jsonify({'status': 'error', 'error_message': 'No active enrollment or already completed'}), 400

# ============================================
# ATTENDANCE SUMMARY & ANALYTICS ENDPOINTS
# ============================================

@app.route('/api/attendance/summary', methods=['GET'])
def api_attendance_summary():
    """
    Get attendance summary with multiple view modes:
    - detailed: Full check-in logs with all metadata
    - subject: Grouped by subject, department, semester, instructor
    - student: Student performance report with attendance rates
    """
    try:
        view_mode = request.args.get('mode', 'detailed')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        department = request.args.get('department')
        semester = request.args.get('semester')
        subject = request.args.get('subject')
        year = request.args.get('year')
        instructor = request.args.get('instructor')
        search = request.args.get('search', '').strip()
        
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Build base query with filters
        base_query = """
            SELECT 
                a.id,
                a.student_id,
                s.name as student_name,
                s.course,
                s.year as student_year,
                a.date,
                a.time,
                a.department,
                a.semester,
                a.year as session_year,
                a.subject,
                a.class_name,
                a.section,
                a.instructor,
                'Present' as status
            FROM attendance a
            JOIN students s ON s.id = a.student_id
            WHERE 1=1
        """
        params = []
        
        if date_from and date_to:
            base_query += " AND a.date BETWEEN ? AND ?"
            params.extend([date_from, date_to])
        elif date_from:
            base_query += " AND a.date >= ?"
            params.append(date_from)
        elif date_to:
            base_query += " AND a.date <= ?"
            params.append(date_to)
            
        if department:
            base_query += " AND a.department = ?"
            params.append(department)
        if semester:
            base_query += " AND a.semester = ?"
            params.append(semester)
        if subject:
            base_query += " AND a.subject = ?"
            params.append(subject)
        if year:
            base_query += " AND a.year = ?"
            params.append(year)
        if instructor:
            base_query += " AND a.instructor = ?"
            params.append(instructor)
        if search:
            base_query += " AND (LOWER(s.name) LIKE ? OR LOWER(a.student_id) LIKE ?)"
            params.extend([f'%{search.lower()}%', f'%{search.lower()}%'])
        
        if view_mode == 'detailed':
            # Detailed logs view
            query = base_query + " ORDER BY a.date DESC, a.time DESC"
            rows = cursor.execute(query, params).fetchall()
            conn.close()
            
            records = []
            for idx, row in enumerate(rows, 1):
                r = dict(row)
                records.append({
                    'index': idx,
                    'student_id': r['student_id'],
                    'student_name': r['student_name'],
                    'course': r['course'],
                    'year': r.get('session_year') or r.get('student_year') or '-',
                    'date': r['date'],
                    'time': r['time'],
                    'department': r.get('department', '-'),
                    'semester': r.get('semester', '-'),
                    'subject': r.get('subject', '-'),
                    'class_name': r.get('class_name', '-'),
                    'section': r.get('section', '-'),
                    'instructor': r.get('instructor', '-'),
                    'status': r['status']
                })
            
            return jsonify({
                'status': 'success',
                'mode': 'detailed',
                'total_records': len(records),
                'records': records
            })
            
        elif view_mode == 'subject':
            # Subject summary view
            query = """
                SELECT 
                    a.department,
                    a.semester,
                    a.year,
                    a.subject,
                    a.instructor,
                    COUNT(DISTINCT a.session_id) as total_sessions,
                    COUNT(*) as present_logs,
                    COUNT(DISTINCT a.student_id) as unique_students,
                    ROUND(
                        COUNT(*) * 100.0 /
                        NULLIF((SELECT COUNT(*) FROM students) * COUNT(DISTINCT a.session_id), 0),
                        1
                    ) as attendance_rate
                FROM attendance a
                JOIN students s ON s.id = a.student_id
                WHERE 1=1
            """
            # Add filters
            if date_from and date_to:
                query += " AND a.date BETWEEN ? AND ?"
            elif date_from:
                query += " AND a.date >= ?"
            elif date_to:
                query += " AND a.date <= ?"
            if department:
                query += " AND a.department = ?"
            if semester:
                query += " AND a.semester = ?"
            if subject:
                query += " AND a.subject = ?"
            if year:
                query += " AND a.year = ?"
            if instructor:
                query += " AND a.instructor = ?"
            if search:
                query += " AND (LOWER(s.name) LIKE ? OR LOWER(a.student_id) LIKE ?)"
                
            query += """ 
                GROUP BY a.department, a.semester, a.year, a.subject, a.instructor
                ORDER BY a.department, a.year, a.semester, a.subject, a.instructor
            """
            
            rows = cursor.execute(query, params).fetchall()
            conn.close()
            
            records = []
            for idx, row in enumerate(rows, 1):
                r = dict(row)
                records.append({
                    'index': idx,
                    'department': r.get('department', '-'),
                    'semester': r.get('semester', '-'),
                    'year': r.get('year', '-'),
                    'subject': r.get('subject', '-'),
                    'instructor': r.get('instructor', '-'),
                    'total_sessions': r['total_sessions'],
                    'present_logs': r['present_logs'],
                    'unique_students': r['unique_students'],
                    'attendance_rate': r.get('attendance_rate', 0)
                })
            
            return jsonify({
                'status': 'success',
                'mode': 'subject',
                'total_subjects': len(records),
                'records': records
            })
            
        elif view_mode == 'student':
            # Student performance view. Filter the sessions first, then retain
            # every enrolled student (including students with no attendance).
            query = """
                WITH filtered_attendance AS (
                    SELECT a.*
                    FROM attendance a
                    WHERE 1=1
            """
            student_params = []
            if date_from and date_to:
                query += " AND a.date BETWEEN ? AND ?"
                student_params.extend([date_from, date_to])
            elif date_from:
                query += " AND a.date >= ?"
                student_params.append(date_from)
            elif date_to:
                query += " AND a.date <= ?"
                student_params.append(date_to)
            if department:
                query += " AND a.department = ?"
                student_params.append(department)
            if semester:
                query += " AND a.semester = ?"
                student_params.append(semester)
            if subject:
                query += " AND a.subject = ?"
                student_params.append(subject)
            if year:
                query += " AND a.year = ?"
                student_params.append(year)
            if instructor:
                query += " AND a.instructor = ?"
                student_params.append(instructor)

            query += """
                ), total AS (
                    SELECT COUNT(DISTINCT session_id) AS total_sessions
                    FROM filtered_attendance
                )
                SELECT 
                    s.id as student_id,
                    s.name as student_name,
                    s.course,
                    s.year as student_year,
                    COUNT(DISTINCT a.session_id) as present_days,
                    total.total_sessions as total_days,
                    ROUND(COUNT(DISTINCT a.session_id) * 100.0 / NULLIF(total.total_sessions, 0), 1) as attendance_rate,
                    MAX(a.date) as last_checkin
                FROM students s
                CROSS JOIN total
                LEFT JOIN filtered_attendance a ON s.id = a.student_id
                WHERE 1=1
            """
            if search:
                query += " AND (LOWER(s.name) LIKE ? OR LOWER(s.id) LIKE ?)"
                student_params.extend([f'%{search.lower()}%', f'%{search.lower()}%'])
            query += " GROUP BY s.id, total.total_sessions ORDER BY attendance_rate DESC, s.name ASC"
            rows = cursor.execute(query, student_params).fetchall()
            
            conn.close()
            
            records = []
            for idx, row in enumerate(rows, 1):
                r = dict(row)
                total_days = r.get('total_days', 1) or 1
                present_days = r.get('present_days', 0) or 0
                absent_days = total_days - present_days
                rate = r.get('attendance_rate', 0) or 0
                
                # Color coding
                if rate >= 80:
                    status = 'good'
                elif rate >= 50:
                    status = 'warning'
                else:
                    status = 'danger'
                
                records.append({
                    'index': idx,
                    'student_id': r['student_id'],
                    'student_name': r['student_name'],
                    'course': r['course'],
                    'year': r.get('student_year') or '-',
                    'present_days': present_days,
                    'absent_days': absent_days,
                    'total_days': total_days,
                    'attendance_rate': rate,
                    'status': status,
                    'last_checkin': r.get('last_checkin', '-')
                })
            
            return jsonify({
                'status': 'success',
                'mode': 'student',
                'total_students': len(records),
                'records': records
            })
        else:
            conn.close()
            return jsonify({'status': 'error', 'error_message': 'Invalid view mode'}), 400
            
    except Exception as e:
        logger.error(f"Attendance summary error: {e}")
        return jsonify({'status': 'error', 'error_message': str(e)}), 500


@app.route('/api/attendance/filters', methods=['GET'])
def api_attendance_filters():
    """Get unique filter values for dropdowns"""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        departments = [row[0] for row in cursor.execute(
            'SELECT DISTINCT department FROM attendance WHERE department IS NOT NULL AND department != "" ORDER BY department'
        ).fetchall()]
        
        semesters = [row[0] for row in cursor.execute(
            'SELECT DISTINCT semester FROM attendance WHERE semester IS NOT NULL AND semester != "" ORDER BY semester'
        ).fetchall()]
        
        subjects = [row[0] for row in cursor.execute(
            'SELECT DISTINCT subject FROM attendance WHERE subject IS NOT NULL AND subject != "" ORDER BY subject'
        ).fetchall()]
        years = [row[0] for row in cursor.execute(
            'SELECT DISTINCT year FROM attendance WHERE year IS NOT NULL AND year != "" ORDER BY year'
        ).fetchall()]
        instructors = [row[0] for row in cursor.execute(
            'SELECT DISTINCT instructor FROM attendance WHERE instructor IS NOT NULL AND instructor != "" ORDER BY instructor'
        ).fetchall()]
        
        conn.close()
        
        return jsonify({
            'status': 'success',
            'departments': departments,
            'semesters': semesters,
            'subjects': subjects,
            'years': years,
            'instructors': instructors
        })
    except Exception as e:
        return jsonify({'status': 'error', 'error_message': str(e)}), 500


@app.route('/api/reports/export', methods=['GET'])
def api_export_report():
    """Export attendance data in various formats based on view mode"""
    try:
        format_type = request.args.get('format', 'csv').lower()
        view_mode = request.args.get('mode', 'detailed')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        department = request.args.get('department')
        semester = request.args.get('semester')
        subject = request.args.get('subject')
        year = request.args.get('year')
        instructor = request.args.get('instructor')
        search = request.args.get('search', '').strip()
        
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Build query based on view mode
        if view_mode == 'detailed':
            query = """
                SELECT 
                    s.id as student_id,
                    s.name as student_name,
                    s.course,
                    s.year as student_year,
                    a.year as year,
                    a.date,
                    a.time,
                    a.department,
                    a.semester,
                    a.subject,
                    a.instructor,
                    'Present' as status
                FROM attendance a
                JOIN students s ON s.id = a.student_id
                WHERE 1=1
            """
            params = []
            if date_from and date_to:
                query += " AND a.date BETWEEN ? AND ?"
                params.extend([date_from, date_to])
            elif date_from:
                query += " AND a.date >= ?"
                params.append(date_from)
            elif date_to:
                query += " AND a.date <= ?"
                params.append(date_to)
            if department:
                query += " AND a.department = ?"
                params.append(department)
            if semester:
                query += " AND a.semester = ?"
                params.append(semester)
            if subject:
                query += " AND a.subject = ?"
                params.append(subject)
            if year:
                query += " AND a.year = ?"
                params.append(year)
            if instructor:
                query += " AND a.instructor = ?"
                params.append(instructor)
            if search:
                query += " AND (LOWER(s.name) LIKE ? OR LOWER(s.id) LIKE ?)"
                params.extend([f'%{search.lower()}%', f'%{search.lower()}%'])
            query += " ORDER BY a.date DESC, a.time DESC"
            
            rows = cursor.execute(query, params).fetchall()
            conn.close()
            
            if not rows:
                return jsonify({'status': 'error', 'error_message': 'No data to export'}), 404
            
            data = [dict(row) for row in rows]
            headers = ['Student ID', 'Student Name', 'Course', 'Year', 'Date', 'Time', 
                      'Department', 'Semester', 'Subject', 'Instructor', 'Status']
            filename = f"attendance_logs_{datetime.now().strftime('%Y%m%d')}"
            
        elif view_mode == 'subject':
            query = """
                SELECT 
                    a.department,
                    a.semester,
                    a.year,
                    a.subject,
                    a.instructor,
                    COUNT(DISTINCT a.session_id) as total_sessions,
                    COUNT(*) as present_logs,
                    COUNT(DISTINCT a.student_id) as unique_students,
                    ROUND(
                        COUNT(*) * 100.0 /
                        NULLIF((SELECT COUNT(*) FROM students) * COUNT(DISTINCT a.session_id), 0),
                        1
                    ) as attendance_rate
                FROM attendance a
                JOIN students s ON s.id = a.student_id
                WHERE 1=1
            """
            params = []
            if date_from and date_to:
                query += " AND a.date BETWEEN ? AND ?"
                params.extend([date_from, date_to])
            elif date_from:
                query += " AND a.date >= ?"
                params.append(date_from)
            elif date_to:
                query += " AND a.date <= ?"
                params.append(date_to)
            if department:
                query += " AND a.department = ?"
                params.append(department)
            if semester:
                query += " AND a.semester = ?"
                params.append(semester)
            if subject:
                query += " AND a.subject = ?"
                params.append(subject)
            if year:
                query += " AND a.year = ?"
                params.append(year)
            if instructor:
                query += " AND a.instructor = ?"
                params.append(instructor)
            query += " GROUP BY a.department, a.semester, a.year, a.subject, a.instructor ORDER BY a.department, a.year, a.semester, a.subject"
            
            rows = cursor.execute(query, params).fetchall()
            conn.close()
            
            if not rows:
                return jsonify({'status': 'error', 'error_message': 'No data to export'}), 404
            
            data = [dict(row) for row in rows]
            headers = ['Department', 'Semester', 'Year', 'Subject', 'Instructor', 'Total Sessions', 
                      'Present Logs', 'Unique Students', 'Attendance Rate (%)']
            filename = f"subject_summary_{datetime.now().strftime('%Y%m%d')}"
            
        elif view_mode == 'student':
            # Keep the export exactly consistent with the Student Performance
            # screen: all selected filters apply and sessions—not dates—are counted.
            query = """
                WITH filtered_attendance AS (
                    SELECT a.*
                    FROM attendance a
                    WHERE 1=1
            """
            params = []
            if date_from and date_to:
                query += " AND a.date BETWEEN ? AND ?"
                params.extend([date_from, date_to])
            elif date_from:
                query += " AND a.date >= ?"
                params.append(date_from)
            elif date_to:
                query += " AND a.date <= ?"
                params.append(date_to)
            if department:
                query += " AND a.department = ?"
                params.append(department)
            if semester:
                query += " AND a.semester = ?"
                params.append(semester)
            if subject:
                query += " AND a.subject = ?"
                params.append(subject)
            if year:
                query += " AND a.year = ?"
                params.append(year)
            if instructor:
                query += " AND a.instructor = ?"
                params.append(instructor)
            query += """
                ), total AS (
                    SELECT COUNT(DISTINCT session_id) AS total_sessions
                    FROM filtered_attendance
                )
                SELECT 
                    s.id as student_id,
                    s.name as student_name,
                    s.course,
                    s.year,
                    COUNT(DISTINCT a.session_id) as present_days,
                    total.total_sessions as total_days,
                    ROUND(COUNT(DISTINCT a.session_id) * 100.0 / NULLIF(total.total_sessions, 0), 1) as attendance_rate,
                    MAX(a.date) as last_checkin
                FROM students s
                CROSS JOIN total
                LEFT JOIN filtered_attendance a ON s.id = a.student_id
                WHERE 1=1
            """
            if search:
                query += " AND (LOWER(s.name) LIKE ? OR LOWER(s.id) LIKE ?)"
                params.extend([f'%{search.lower()}%', f'%{search.lower()}%'])
            query += " GROUP BY s.id, total.total_sessions ORDER BY attendance_rate DESC, s.name ASC"
            rows = cursor.execute(query, params).fetchall()
            conn.close()
            
            if not rows:
                return jsonify({'status': 'error', 'error_message': 'No data to export'}), 404
            
            data = []
            for row in rows:
                r = dict(row)
                total_days = r.get('total_days', 1) or 1
                present_days = r.get('present_days', 0) or 0
                data.append({
                    'student_id': r['student_id'],
                    'student_name': r['student_name'],
                    'course': r['course'],
                    'year': r['year'],
                    'present_days': present_days,
                    'absent_days': total_days - present_days,
                    'total_days': total_days,
                    'attendance_rate': r.get('attendance_rate', 0),
                    'last_checkin': r.get('last_checkin', '-')
                })
            headers = ['Student ID', 'Student Name', 'Course', 'Year', 'Present Days', 
                      'Absent Days', 'Total Days', 'Attendance Rate (%)', 'Last Check-in']
            filename = f"student_performance_{datetime.now().strftime('%Y%m%d')}"
        else:
            conn.close()
            return jsonify({'status': 'error', 'error_message': 'Invalid view mode'}), 400
        
        # Export based on format
        if format_type == 'csv':
            return export_as_csv(data, headers, filename)
        elif format_type == 'excel':
            return export_as_excel(data, headers, filename, view_mode)
        elif format_type == 'json':
            return export_as_json(data, filename)
        else:
            return jsonify({'status': 'error', 'error_message': f'Unsupported format: {format_type}'}), 400
            
    except Exception as e:
        logger.error(f"Export error: {e}")
        return jsonify({'status': 'error', 'error_message': str(e)}), 500


def export_as_csv(data, headers, filename):
    """Export data as CSV with UTF-8 BOM encoding for Excel compatibility"""
    import csv
    from io import StringIO
    
    output = StringIO()
    # Add UTF-8 BOM
    output.write('\uFEFF')
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(headers)
    
    # Write data
    for row in data:
        writer.writerow([row.get(h.replace(' ', '_').replace('(', '').replace(')', '').replace('%', '').strip('_').lower(), '') for h in headers])
    
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename={filename}.csv'}
    )


def export_as_excel(data, headers, filename, view_mode):
    """Export data as Excel file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        for row_idx, row in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                key = header.replace(' ', '_').lower()
                value = row.get(key, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # Color code attendance rate
                if 'attendance_rate' in key and value:
                    try:
                        rate = float(value)
                        if rate >= 80:
                            cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                        elif rate >= 50:
                            cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                    except:
                        pass
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 35)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_ws['A1'] = f"Attendance Report - {view_mode.title()}"
        summary_ws['A1'].font = Font(size=14, bold=True)
        summary_ws.merge_cells('A1:B1')
        
        summary_ws['A3'] = "Generated On"
        summary_ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        summary_ws['A4'] = "Total Records"
        summary_ws['B4'] = len(data)
        
        summary_ws['A5'] = "Report Type"
        summary_ws['B5'] = view_mode.title()
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}.xlsx'}
        )
    except ImportError:
        return jsonify({'status': 'error', 'error_message': 'openpyxl not installed. Please install: pip install openpyxl'}), 500


def export_as_json(data, filename):
    """Export as JSON"""
    return jsonify({
        'status': 'success',
        'export_date': datetime.now().isoformat(),
        'total_records': len(data),
        'data': data
    })


# --- Dashboard Stats API ---
@app.route('/api/stats', methods=['GET'])
def api_get_stats():
    try:
        students = database.get_all_students()
        total_students = len(students)
        
        today_str = datetime.today().strftime('%Y-%m-%d')
        
        conn = database.get_db_connection()
        present_count = conn.execute(
            'SELECT COUNT(DISTINCT student_id) FROM attendance WHERE date = ?', (today_str,)
        ).fetchone()[0]
        
        recent_rows = conn.execute('''
            SELECT s.name, a.time, a.date, a.subject, a.department
            FROM attendance a
            JOIN students s ON s.id = a.student_id
            ORDER BY a.id DESC LIMIT 5
        ''').fetchall()
        conn.close()
        
        recent_activities = []
        for r in recent_rows:
            t_parsed = datetime.strptime(r['time'], '%H:%M:%S')
            t_formatted = t_parsed.strftime('%I:%M %p')
            detail = r['subject'] or ''
            if r['department']:
                detail = f"{r['department']} - {detail}" if detail else r['department']
            recent_activities.append({
                'time': t_formatted,
                'name': r['name'],
                'detail': detail,
                'status': 'Present'
            })
            
        absent_count = max(0, total_students - present_count)
        
        return jsonify({
            'status': 'success',
            'total_students': total_students,
            'present_today': present_count,
            'absent_today': absent_count,
            'recent_activity': recent_activities
        })
    except Exception as e:
        return jsonify({'status': 'error', 'error_message': str(e)}), 500

# --- Attendance Records Query API ---
@app.route('/api/attendance', methods=['GET'])
def api_query_attendance():
    try:
        date_filter = request.args.get('date', '').strip() or None
        date_to = request.args.get('date_to', '').strip() or None
        search_query = request.args.get('search', '')
        
        records = database.get_attendance_records(date_filter, date_to, search_query)
        return jsonify({'status': 'success', 'records': records})
    except Exception as e:
        return jsonify({'status': 'error', 'error_message': str(e)}), 500

# --- Webcam Live Video Session APIs ---
@app.route('/api/live/start', methods=['GET', 'POST'])
def api_start_camera():
    success = camera.start()
    if success:
        return jsonify({'status': 'success', 'message': 'Camera session started'})
    else:
        return jsonify({'status': 'error', 'error_message': 'Failed to initialize webcam'}), 500

@app.route('/api/live/stop', methods=['GET', 'POST'])
def api_stop_camera():
    try:
        data = request.get_json(silent=True) or {}
        save = data.get('save', True)
        session_meta = {
            'department':  data.get('department', ''),
            'semester':    data.get('semester', ''),
            'year':        data.get('year', ''),
            'subject':     data.get('subject', ''),
            'instructor':  data.get('instructor', ''),
            'class_name':  data.get('class_name', ''),
            'section':     data.get('section', ''),
            'session_id':  uuid.uuid4().hex,
        }
        if save:
            required = ('department', 'semester', 'year', 'subject', 'instructor')
            missing = [field.replace('_', ' ').title() for field in required if not str(session_meta[field]).strip()]
            if missing:
                return jsonify({'status': 'error', 'error_message': f"Complete session details before saving: {', '.join(missing)}."}), 400
        if save:
            camera.save_session_to_database(session_meta)
        camera.stop()
        msg = 'Session saved and stopped.' if save else 'Session discarded and stopped.'
        return jsonify({'status': 'success', 'message': msg})
    except Exception as e:
        logger.error(f"Error in api_stop_camera: {e}")
        camera.stop()
        return jsonify({'status': 'error', 'error_message': str(e)}), 500


@app.route('/api/live/reset', methods=['POST'])
def api_reset_camera_session():
    camera.reset_session()
    return jsonify({'status': 'success', 'message': 'Active session counters reset'})

@app.route('/api/live/status', methods=['GET'])
def api_get_live_status():
    students = database.get_all_students()
    
    attendance_list = []
    for s in students:
        is_detected = s['id'] in camera.session_attendance
        attendance_list.append({
            'id': s['id'],
            'name': s['name'],
            'status': 'Present' if is_detected else 'Absent',
            'time': camera.session_attendance[s['id']] if is_detected else '-'
        })
        
    return jsonify({
        'status': 'success',
        'is_running': camera.is_running,
        'faces_detected': camera.faces_detected,
        'recognized_count': len(camera.session_attendance),
        'unknown_count': camera.unknown_faces,
        'attendance_list': attendance_list,
        'total_students': len(students)
    })

# Generator function for streaming video
def gen_frames():
    while True:
        if not camera.is_running or camera.last_frame is None:
            time.sleep(0.1)
            continue
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + camera.last_frame + b'\r\n')
        time.sleep(0.04)

@app.route('/video_feed')
def video_feed():
    return Response(gen_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
